"""normalize_canvass.py — Washington County Clerk canvass crosstabs → canonical tidy long file.

Input:  raw/<year>/<month>/…  (verbatim mirror of the Washington County Clerk's
        published precinct-level results files, outpost.washco.utah.gov — see
        sources.csv for the complete byte-verified catalog).
Output: washco_results_long.csv — one row per precinct × candidate(-column) per
        election, SLCo-schema-compatible columns:
        year,election_type,source_file,sheet,contest,vote_for,precinct,candidate,
        votes,suppressed,vote_method,times_cast,registered_voters

Format families (verified from file bodies, 2026-07-20):
  E1  export crosstab: 3 header rows (contest / party-code / candidate), leading
      meta cols COUNTY NUMBER, PRECINCT CODE, PRECINCT NAME, REGISTERED VOTERS
      TOTAL, BALLOTS CAST TOTAL, BALLOTS CAST BLANK; data rows per precinct;
      final row PRECINCT CODE=ZZZ "COUNTY TOTALS" = the certified grand total.
      2018/11 is the same layout inside an .xlsx worksheet.
  P   partisan-primary variant of E1: extra per-party REGISTERED VOTERS - <party>
      / BALLOTS CAST - <party> meta columns (2020-03, 2020-06 "SOVC" CSVs,
      2023-09, 2024-03 DPP, 2024-06).
  E2  2024/11 variant: NO registered-voters/ballots-cast meta columns at all
      (COUNTY NUMBER, PRECINCT CODE, PRECINCT NAME only).

Faithfulness rules:
  - Contest, candidate, and precinct-name strings are VERBATIM (incl. the
    2021-11 file's jurisdiction-suffixed contest names and pseudo-candidate
    columns such as "OVER VOTES"/"UNDER VOTES"/"Write-in"/"WITHDREW" — they are
    part of the published canvass and are retained here; the derived by-contest
    layer filters non-candidate columns).
  - Zero cells are emitted: the crosstab prints every precinct row for every
    contest column, including precincts outside a contest's jurisdiction. The
    long file is an exact tidy transform of the published crosstab. Consumers
    should not read "precinct appears" as "precinct is in the jurisdiction" —
    build_elections.py counts n_precincts as NONZERO precincts.
  - The ZZZ/"COUNTY TOTALS" row is NOT emitted as a precinct; it is used as the
    per-column reconciliation gate (sum of precinct rows must equal it exactly)
    and the build FAILS if any column mismatches.
  - No suppression markers exist in any machine-readable file (verified: zero
    '*' cells) → suppressed=False throughout. The county's only redacted
    publication is the Jun-2026 precinct PDF, which is not loaded (see
    VERIFICATION.md).

DERIVED + idempotent. Never hand-edit the output; rerun this.
"""
import csv
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, "raw")
OUT = os.path.join(HERE, "washco_results_long.csv")

# (year, election_type, relpath under raw/) — machine-readable precinct files only.
# PDF-only elections (2018-06 primary, 2022-06 primary, 2026-06 primary) are
# catalogued in sources.csv and ledgered in VERIFICATION.md as machine-readable
# gaps — never fabricated from PDFs here.
FILES = [
    (2018, "general",             "2018/11/201811-election-results-by-precinct.xlsx"),
    (2019, "municipal general",   "2019/11/2019-general-municipal-export.csv"),
    (2020, "presidential primary","2020/reports/washco-election-sovc-202003.CSV"),
    (2020, "primary",             "2020/reports/washco-election-sovc-202006.CSV"),
    (2020, "general",             "2020/11/washco-election-20201103-results-export.csv"),
    (2021, "municipal primary",   "2021/08/washco_elections_20210810_OFFICIAL_municipal-primary.csv"),
    (2021, "municipal general",   "2021/11/washco-election-20211102-results-export.csv"),
    (2022, "general",             "2022/11/washco-election-20221122-results-export.csv"),
    (2023, "municipal primary",   "2023/09/washco-election-202309-results-export.csv"),
    (2023, "municipal general",   "2023/11/washco-election-20231121-results-export.csv"),
    (2024, "presidential primary","2024/03/washco-elections-202403-dpp-official-export.csv"),
    (2024, "primary",             "2024/06/washco-election-20240625-results-export.csv"),
    (2024, "general",             "2024/11/washco-election-20241105-results-export.csv"),
    (2025, "municipal primary",   "2025/08/washco-election-20250812-results-export.csv"),
    (2025, "municipal general",   "2025/11/washco-election-20251104-results-export.csv"),
]

META_EXACT = {"COUNTY NUMBER", "PRECINCT CODE", "PRECINCT NAME"}
META_PREFIX = ("REGISTERED VOTERS", "BALLOTS CAST")

# Source-internal contradictions, kept VERBATIM (cardinal rule: city/county-faithful
# values are never overwritten). Each entry is a (source relpath, contest, candidate)
# whose precinct-row sum disagrees with the file's own certified ZZZ COUNTY TOTALS
# row. Ledgered in VERIFICATION.md; anything NOT listed here still fails the build.
#   2019 Dammeron Valley Fire SSD (single-precinct contest, CODAV): precinct row
#   prints STEWART 132 / Write-in 122, certified totals row prints 127 / 121
#   (THOMAS 184 agrees). Both figures are the county's own publication; no second
#   source exists (no 2019 precinct PDF was published). Non-municipal contest —
#   never reaches the derived/gov.db layer.
KNOWN_SOURCE_DISCREPANCIES = {
    ("2019/11/2019-general-municipal-export.csv", "Dammeron Valley Fire SSD", "DENISE STEWART"),
    ("2019/11/2019-general-municipal-export.csv", "Dammeron Valley Fire SSD", "Write-in"),
}


def read_rows(path):
    """Return (rows, sheet_name). CSV → list of rows; XLSX → first sheet."""
    if path.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [["" if c is None else str(c) for c in r]
                for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows, ws.title
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        return list(csv.reader(f)), ""


def is_meta(contest_hdr):
    cu = " ".join((contest_hdr or "").split()).upper()
    return (not cu) or cu in META_EXACT or cu.startswith(META_PREFIX)


def parse_file(year, etype, relpath):
    """Yield long rows; verify every candidate column against the ZZZ totals row."""
    path = os.path.join(RAW, relpath)
    rows, sheet = read_rows(path)
    hdr_contest, hdr_cand = rows[0], rows[2]
    ncol = len(hdr_contest)

    def col(name):
        for i, c in enumerate(hdr_contest):
            if " ".join((c or "").split()).upper() == name:
                return i
        return None

    code_col, name_col = col("PRECINCT CODE"), col("PRECINCT NAME")
    rv_col, bc_col = col("REGISTERED VOTERS TOTAL"), col("BALLOTS CAST TOTAL")
    assert code_col is not None and name_col is not None, relpath

    cand_cols = [i for i in range(ncol)
                 if not is_meta(hdr_contest[i])
                 and i < len(hdr_cand) and (hdr_cand[i] or "").strip()]

    data, totals_row = [], None
    for r in rows[3:]:
        if len(r) <= name_col:
            continue
        code = (r[code_col] or "").strip()
        if not code:
            continue
        if code == "ZZZ":
            totals_row = r
            continue
        data.append(r)

    # per-precinct-name uniqueness (verified 2026-07-20 for every file; keep the gate)
    names = [(r[name_col] or "").strip() for r in data]
    assert len(names) == len(set(names)), f"duplicate precinct names in {relpath}"

    # reconciliation gate: column sums == certified COUNTY TOTALS row
    assert totals_row is not None, f"no ZZZ COUNTY TOTALS row in {relpath}"
    bad = []
    for i in cand_cols:
        s = sum(int(float(r[i])) for r in data if i < len(r) and (r[i] or "").strip())
        t = int(float(totals_row[i])) if i < len(totals_row) and (totals_row[i] or "").strip() else 0
        if s != t:
            contest = " ".join((hdr_contest[i] or "").split())
            cand = " ".join((hdr_cand[i] or "").split())
            if (relpath, contest, cand) in KNOWN_SOURCE_DISCREPANCIES:
                print(f"  KNOWN source discrepancy (kept verbatim) {relpath}: "
                      f"{contest} / {cand}: precinct-sum {s} vs ZZZ {t}")
            else:
                bad.append((contest, cand, s, t))
    if bad:
        for contest, cand, s, t in bad[:10]:
            print(f"  MISMATCH {relpath}: {contest} / {cand}: precinct-sum {s} != ZZZ {t}")
        raise SystemExit(f"reconciliation FAILED for {relpath} ({len(bad)} columns)")

    src = os.path.basename(relpath)
    out = []
    for r in data:
        precinct = (r[name_col] or "").strip()
        rv = (r[rv_col] or "").strip() if rv_col is not None and rv_col < len(r) else ""
        bc = (r[bc_col] or "").strip() if bc_col is not None and bc_col < len(r) else ""
        for i in cand_cols:
            v = (r[i] or "").strip() if i < len(r) else ""
            out.append({
                "year": year, "election_type": etype, "source_file": src,
                "sheet": sheet,
                "contest": " ".join((hdr_contest[i] or "").split()),
                "vote_for": "",
                "precinct": precinct,
                "candidate": " ".join((hdr_cand[i] or "").split()),
                "votes": int(float(v)) if v else 0,
                "suppressed": "False",
                "vote_method": "Total",
                "times_cast": bc, "registered_voters": rv,
            })
    print(f"  OK {relpath}: {len(data)} precincts x {len(cand_cols)} candidate cols "
          f"= {len(out)} rows; all {len(cand_cols)} columns reconcile to COUNTY TOTALS")
    return out


def main():
    cols = ["year", "election_type", "source_file", "sheet", "contest", "vote_for",
            "precinct", "candidate", "votes", "suppressed", "vote_method",
            "times_cast", "registered_voters"]
    all_rows = []
    for year, etype, relpath in FILES:
        all_rows.extend(parse_file(year, etype, relpath))
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(all_rows)
    print(f"Wrote {OUT}: {len(all_rows)} rows from {len(FILES)} canvass files")


if __name__ == "__main__":
    sys.exit(main())
