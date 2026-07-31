"""build_elections.py — build the canonical Utah County election canvass layer.

Utah County (FIPS 49049) publishes its canvass through the Clerk's elections app
(https://vote.utahcounty.gov/results/{year}); the raw files are mirrored verbatim in
raw/ (provenance: sources.csv — byte-verified URLs + sha256). This script parses them
into:

  utah_county_results_long.csv       CANONICAL tidy long — one row per precinct (or
                                     countywide, where the county published no precinct
                                     grain) × contest × candidate. Column-compatible
                                     superset of the SLCo model
                                     (salt_lake_county/elections/slco_municipal_results_long.csv):
                                     the 13 SLCo columns first, then party / grain /
                                     extraction.
  election_results_by_contest.csv    DERIVED — one row per contest × candidate, the 14
                                     loader columns scripts/build_cities_db.py
                                     load_election_result() reads, plus rcv /
                                     rcv_final_winner / official_total (extra columns
                                     are ignored by the loader). Municipal council/mayor
                                     contests (every Utah County municipality;
                                     jurisdiction_slug set for held cities lehi / provo /
                                     orem / vineyard) + Utah County county offices
                                     (jurisdiction_slug='utah_county').
  rcv/rcv_contests.csv               The RCV registry (written by this script from
                                     rcv/rounds/*.json + the county summary PDFs).
  rcv/rounds/*.csv                   Flattened round-by-round tables from the rcvis.com
                                     JSON blobs archived in rcv/rounds/*.json.

DERIVED + idempotent: rerun after adding raws. Never hand-edit outputs.

Cardinal rules honored here:
  * suppressed cells ('-' in the county files) stay suppressed — emitted with votes=''
    and suppressed=True, never imputed;
  * candidate names, contest names, precinct labels are verbatim as printed
    (including "(WITHDREW)"/"(DISQUALIFIED)" suffixes and merged-precinct labels like
    "AF13 & AF14");
  * COUNTY TOTALS rows are the county's own rollup — captured as official contest
    totals (they INCLUDE the suppressed cells), excluded from the precinct rows, and
    reconciled against the precinct sums at build time;
  * RCV contests are never presented in SOVC first-choice order as a final result —
    rank-position contests ("... 2nd Choice") stay out of the by-contest layer, rcv
    contests are flagged, and final winners come from the county's own RCV tabulations
    (rcvis.com, county-linked) recorded in rcv/.
"""
import csv
import glob
import json
import os
import re
import sys
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, "raw")
TEXT = os.path.join(RAW, "text")
OCR19 = os.path.join(RAW, "text", "ocr_2019_general")
RCV_DIR = os.path.join(HERE, "rcv")
LONG_OUT = os.path.join(HERE, "utah_county_results_long.csv")
BYC_OUT = os.path.join(HERE, "election_results_by_contest.csv")

# ---------------------------------------------------------------------------
# source registry: (year, election_type, filename, parser, note)
# Only files parsed into the long layer appear here; everything acquired is in
# sources.csv (catalog + verification-only + mislabeled files included).
# ---------------------------------------------------------------------------
PARSED_SOURCES = [
    (2016, "general",              "2016_General_Results_Summary_997e775ff5.pdf",  "gems_summary"),
    (2016, "regular primary",      "2016_Primary_Results_Summary_f171493cc7.pdf",  "gems_summary"),
    (2017, "municipal general",    "2017_General_Election_Summary_0e2231c9e1.pdf", "evs_summary"),
    (2017, "municipal primary",    "2017_Primary_Lehi_Summary_97c19ea2ea.pdf",     "evs_summary"),
    (2018, "regular primary",      "18_P_SOVC_By_Precinct_P_2018_Suppressed_d4d32a6067.xlsx", "xlsx_sovc"),
    (2018, "general",              "18_G_SOVC_By_Precinct_G_2018_Suppressed_f8b6c0ead9.xlsx", "xlsx_sovc"),
    (2019, "municipal primary",    "19_P_19_Primary_SOVC_suppressed_93de48c7ac.CSV", "wide_csv"),
    (2019, "municipal general",    "19_G_Countywide_Precinct_Official_Suppressed_c07b072cdf.pdf", "ocr_precinct"),
    (2020, "presidential primary", "20_P_2020_Presidential_Primary_SOVC_final_SUPPRESSED_472420ab89.csv", "wide_csv"),
    (2020, "regular primary",      "2020_Primary_SOVC_suppressed_a96f91f7cb.CSV",  "wide_csv"),
    (2020, "general",              "20_G_2020_General_SOVC_SUPPRESSED_2677d8223a.csv", "wide_csv"),
    (2021, "municipal primary",    "21_PP_2021_Primary_Statement_of_Votes_Cast_SUPPRESSED_bd47a35ddf.csv", "wide_csv"),
    (2021, "municipal general",    "21_G_Countywide_SOVC_suppressed_1b85ad469d.csv", "wide_csv"),
    (2022, "regular primary",      "2022_Primary_PDF_1c718e2068.pdf",              "ew_summary"),
    (2022, "general",              "22_G_sovc_SUPPRESSED_c752b2b805.csv",          "wide_csv"),
    (2023, "municipal primary",    "23_P_SOV_Cs_suppressed_1907fb1cba.pdf",        "ew_precinct"),
    (2023, "municipal general",    "2023_General_voting_results_be47c5636c.pdf",   "ew_summary"),
    (2024, "presidential primary", "24_PP_sovc_suppressed_24_Prez_Primary_24488c59cf.csv", "wide_csv"),
    (2024, "regular primary",      "24_P_SOVC_suppressed_small_precincts_41eef5de38.csv", "wide_csv"),
    (2024, "general",              "2024_General_SOVC_FINAL_9d0c1e4b30.csv",       "wide_csv"),
    (2025, "municipal primary",    "2025_Primary_SOVC_suppressed_4bc086dabf.csv",  "wide_csv"),
    (2025, "municipal general",    "SOVC_Simple_Redacted_7a5eddcaf2.csv",          "wide_csv"),
    (2026, "regular primary",      "Precinct_Summary_OFFICIAL_RESULTS_1_69c1ce8468.pdf", "ew_precinct"),
]

# countywide summary PDFs used for reconciliation of the precinct-grain years
# (year, election_type, text filename, format, official?)
# official=False marks the county's interim/unofficial reports (in-body header
# "UNOFFICIAL RESULTS") — expected to run slightly LOW vs the certified SOVC.
RECON_SUMMARIES = [
    (2018, "general",              "2018_General_Results_PDF_f84a6d041b.txt", "evs", True),
    (2018, "regular primary",      "2018_Primary_PDF_fbadf156d7.txt", "evs", True),
    (2019, "municipal general",    "2019_General_Results_PDF_a69d246ddc.txt", "ew", True),
    (2019, "municipal primary",    "2019_Primary_Results_PDF_dba3744ad0.txt", "ew", True),
    (2020, "general",              "2020_General_PDF_3be636af10.txt", "ew", True),
    (2020, "regular primary",      "2020_Primary_PDF_b759ce25d5.txt", "ew", True),
    (2020, "presidential primary", "2020_Presidential_Primary_PDF_2cc741ca9c.txt", "ew", True),
    (2021, "municipal general",    "2021_General_PDF_4d36475691.txt", "ew", True),
    (2021, "municipal primary",    "2021_Primary_PDF_e05a1d3833.txt", "ew", False),
    (2023, "municipal primary",    "2023_Primary_voting_results_30a0ba993f.txt", "ew", True),
    (2024, "general",              "Summary_Results_FINAL_ec1e91cd43.txt", "ew", True),
    (2025, "municipal general",    "OFFICIAL_Countywide_Results_11_17_f09d22f26a.txt", "ew", True),
    (2025, "municipal primary",    "Countywide_Summary_Results_Official_99ee333134.txt", "ew", True),
]

META_TOP = {
    "PRECINCT NAME", "COUNTY NUMBER", "PRECINCT CODE",
    "REGISTERED VOTERS TOTAL", "BALLOTS CAST TOTAL", "BALLOTS CAST BLANK",
    "BALLOTS CAST - REPUBLICAN", "BALLOTS CAST - DEMOCRATIC",
    "BALLOTS CAST - NONPARTISAN",
    "REGISTERED VOTERS - DEMOCRATIC", "REGISTERED VOTERS - REPUBLICAN",
    "REGISTERED VOTERS - NONPARTISAN",
}
PARTY_TOKENS = {"REP", "DEM", "NON", "CON", "LIB", "IAP", "UUP", "GRN", "UNA",
                "UU", "INA", "UNF", "GLC", "AME", "IAM"}
TOTALS_LABELS = {"COUNTY TOTALS", "COUNTYWIDE", "ELECTIONWIDE", "TOTALS", "TOTAL"}

# Ballot-accounting lines the county prints in candidate position — kept verbatim in
# the LONG file, excluded from candidate ranking in the by-contest layer (exact match,
# the SLCo METHOD_LABELS discipline).
PSEUDO_CANDIDATES = {"OVER VOTES", "UNDER VOTES", "OVERVOTES", "UNDERVOTES",
                     "CONTEST TOTALS", "TOTAL VOTES CAST", "TOTAL VOTES",
                     "BALLOTS CAST", "VOTERS", "TIMES CAST"}


def n_or_blank(v):
    v = (v or "").strip().replace(",", "")
    if v in ("", "-"):
        return ""
    try:
        return str(int(float(v)))
    except ValueError:
        return ""


class Row(dict):
    pass


def emit(rows, year, etype, source, sheet, contest, vote_for, precinct, candidate,
         votes, suppressed, times_cast="", registered="", party="", grain="precinct",
         extraction="csv"):
    rows.append({
        "year": year, "election_type": etype, "source_file": source, "sheet": sheet,
        "contest": " ".join(str(contest).split()),
        "vote_for": vote_for, "precinct": precinct,
        "candidate": " ".join(str(candidate).split()),
        "votes": votes, "suppressed": suppressed, "vote_method": "Total",
        "times_cast": times_cast, "registered_voters": registered,
        "party": party, "grain": grain, "extraction": extraction,
    })


# ---------------------------------------------------------------------------
# Parser 1: wide matrix CSV (2019P..2025G) — 2 or 3 header rows, precinct rows,
# trailing COUNTY TOTALS rollup.
# ---------------------------------------------------------------------------
def parse_wide_csv(path, year, etype):
    fname = os.path.basename(path)
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        raw = [r for r in csv.reader(f)]
    top = [" ".join(c.split()) for c in raw[0]]
    # carry contest names forward across blank header cells (defensive)
    for i in range(1, len(top)):
        if not top[i] and i > 0 and top[i - 1] and top[i - 1] not in META_TOP:
            top[i] = top[i - 1]
    # detect party row
    r1 = [c.strip() for c in raw[1]]
    data_cols = [i for i, c in enumerate(top) if c and c not in META_TOP]
    r1_vals = {r1[i] for i in data_cols if i < len(r1) and r1[i]}
    has_party = bool(r1_vals) and r1_vals <= PARTY_TOKENS
    cand_row_i = 2 if has_party else 1
    cands = [c.strip() for c in raw[cand_row_i]]
    party = r1 if has_party else [""] * len(top)

    def col(name):
        return top.index(name) if name in top else None

    c_name, c_reg, c_cast = col("PRECINCT NAME"), col("REGISTERED VOTERS TOTAL"), col("BALLOTS CAST TOTAL")
    if c_cast is None:
        c_cast = col("BALLOTS CAST - REPUBLICAN")   # 2020 regular primary (REP-only ballot)
    rows, totals = [], {}
    for r in raw[cand_row_i + 1:]:
        if not any(x.strip() for x in r):
            continue
        name = r[c_name].strip() if c_name is not None and c_name < len(r) else ""
        if not name:
            continue
        is_total = name.upper() in TOTALS_LABELS
        reg = n_or_blank(r[c_reg]) if c_reg is not None and c_reg < len(r) else ""
        cast = n_or_blank(r[c_cast]) if c_cast is not None and c_cast < len(r) else ""
        for i in data_cols:
            if i >= len(r):
                continue
            cell = r[i].strip()
            cand = cands[i] if i < len(cands) else ""
            if not cand:
                continue
            if is_total:
                if cell not in ("", "-"):
                    totals[(top[i], cand)] = int(cell.replace(",", ""))
                continue
            if cell == "-":
                emit(rows, year, etype, fname, "", top[i], "", name, cand, "", True,
                     cast, reg, party[i] if i < len(party) else "")
            elif cell != "":
                v = int(cell.replace(",", ""))
                emit(rows, year, etype, fname, "", top[i], "", name, cand, v, False,
                     cast, reg, party[i] if i < len(party) else "")
    # participation filter: keep a (precinct, contest) block only if any cell is
    # nonzero or suppressed (the county matrix prints 0 for off-ballot precincts —
    # an all-zero block is indistinguishable from "not on this precinct's ballot").
    by_pc = defaultdict(list)
    for r in rows:
        by_pc[(r["precinct"], r["contest"])].append(r)
    kept = []
    for _, rs in by_pc.items():
        if any(r["suppressed"] or (r["votes"] != "" and int(r["votes"]) > 0) for r in rs):
            kept.extend(rs)
    return kept, totals


# ---------------------------------------------------------------------------
# Parser 2: 2018 SOVC xlsx — one contest per sheet.
# ---------------------------------------------------------------------------
def parse_xlsx_sovc(path, year, etype):
    import openpyxl
    fname = os.path.basename(path)
    wb = openpyxl.load_workbook(path, read_only=True)
    rows, totals = [], {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        grid = [[("" if v is None else str(v)) for v in r]
                for r in ws.iter_rows(values_only=True)]
        if len(grid) < 6:
            continue
        title = " ".join(grid[1][0].split()) if grid[1] else ""
        m = re.match(r"(.*?)\s*\(Vote for\s+(\d+)\)", title)
        if not m:      # Sheet1 = turnout page, no contest
            continue
        contest, vote_for = m.group(1).strip(), m.group(2)
        # header row: the row whose col A == 'Precinct'
        hi = next((i for i, r in enumerate(grid) if r and r[0].strip() == "Precinct"), None)
        if hi is None:
            continue
        hdr = grid[hi]
        # candidate columns start after the SECOND 'Precinct' column
        pcols = [i for i, c in enumerate(hdr) if c.strip() == "Precinct"]
        start = (pcols[1] if len(pcols) > 1 else pcols[0]) + 1
        cand_cols = []
        for i in range(start, len(hdr)):
            h = hdr[i].strip()
            if not h:
                continue
            if h.replace("\n", " ").strip() == "Total Votes":
                continue
            nm = " ".join(h.split("\n")[0].split())
            pm = re.search(r"\(([A-Z]{2,4})\)", h)
            cand_cols.append((i, nm, pm.group(1) if pm else "",
                              "Qualified Write-in" in h))
        c_cast = next((i for i, c in enumerate(hdr) if "Times Cast" in c), None)
        c_reg = next((i for i, c in enumerate(hdr) if "Registered" in c), None)
        for r in grid[hi + 1:]:
            if not r or not r[0].strip():
                continue
            name = r[0].strip()
            if name in ("Precinct", "County", "Utah") or name.startswith("Page:"):
                continue
            if name == "County - Total":       # the workbook's official rollup
                for i, cand, pty, _ in cand_cols:
                    v = n_or_blank(r[i] if i < len(r) else "")
                    if v != "":
                        totals[(contest, cand)] = int(v)
                break   # the primary workbook repeats the precinct list once per
                        # district grouping (County/Congressional/State House/...)
                        # with identical values — parse only the County block
            if (name.upper() in TOTALS_LABELS or name.startswith("Total")
                    or name.endswith("- Total") or name.startswith("Cumulative")):
                continue                        # 'Utah - Total' / 'Cumulative' rollups
            cast = n_or_blank(r[c_cast]) if c_cast is not None and c_cast < len(r) else ""
            reg = n_or_blank(r[c_reg]) if c_reg is not None and c_reg < len(r) else ""
            wrote = []
            for i, cand, pty, is_wi in cand_cols:
                cell = (r[i] if i < len(r) else "").strip()
                if cell == "-":
                    wrote.append((cand, "", True, pty))
                elif cell != "":
                    try:
                        wrote.append((cand, int(float(cell.replace(",", ""))), False, pty))
                    except ValueError:
                        pass
            if any(s or (v != "" and v > 0) for _, v, s, _ in wrote):
                for cand, v, s, pty in wrote:
                    emit(rows, year, etype, fname, sn, contest, vote_for, name, cand,
                         v, s, cast, reg, pty, extraction="xlsx")
    return rows, totals


# ---------------------------------------------------------------------------
# Parser 3: Electionware precinct-summary PDF text (2023 primary, 2026 primary,
# and the 2019 general OCR pages). One or more precincts per page; page header
# then precinct code line, STATISTICS block, contest blocks.
# ---------------------------------------------------------------------------
STAT_RE = re.compile(r"^(Registered Voters|Ballots Cast|Voter Turnout|Times Cast|Statistics|STATISTICS|TOTAL$)")
CAND_RE = re.compile(r"^(.{2,}?)\s\s+([\d,]+)(?:\s+([\d.]+)\s*%)?\s*$")
# OCR collapses runs of spaces to one; allow single-space separation there, and
# tolerate OCR junk inside the trailing percent token ("§2.42%").
OCR_CAND_RE = re.compile(r"^(.{2,}?)\s+([\d,]+)(?:\s+(\S*[\d.]\S*%))?\s*$")

# Documented OCR corrections, each visually verified against the raw PDF page
# (pdftoppm render + manual read). Keyed by (precinct, contest, candidate).
# The tesseract pass dropped the SHIPLEY line on report page 189 (PR33) —
# printed values: SHIPLEY 58 / MOSS 47 / Total 105.
OCR_PATCHES_2019G = {
    ("PR33", "Provo City Council - City Wide II", "DAVID SHIPLEY"): 58,
}
# Precinct-label OCR repairs that the generic 2-letter-prefix rule cannot make,
# each derived from the report's alphabetical page order (PROO sits between PR08
# and PR10) or unambiguous city context (Vineyard).
OCR_PRECINCT_PATCHES = {"PROO": "PR09", "VVIO2": "VI02", "V1L01S2": "VI01S2"}
PCT_ONLY_RE = re.compile(r"^[\d.,]+%?$")
PRECINCT_RE = re.compile(r"^[0-9A-Z]{2,10}(?::UN|S)?(?:\s*&\s*[0-9A-Z]{2,10}(?::UN|S)?)*$")


def _fix_ocr_candidate(c):
    """Repair systematic tesseract confusions in candidate names (verified against
    the certified summary: '$'->'S', 'S$'->'S', '!'->'I'). Source-faithful — the
    printed names have no $/!."""
    c = re.sub(r"\bS\$", "S", c)
    c = c.replace("$", "S").replace("!", "I")
    return c


def _fix_ocr_precinct(tok):
    # OCR confuses 0/O inside precinct codes. Utah County precinct codes are a
    # 2-letter city prefix + numeric tail (optional S/S2/:UN suffix), so fix the
    # tail after the first two letters.
    def fix(t):
        if t in OCR_PRECINCT_PATCHES:
            return OCR_PRECINCT_PATCHES[t]
        m = re.match(r"^([A-Z]{2})([0-9OQ]+[A-Z0-9]?(?::UN)?)$", t)
        if m:
            return m.group(1) + m.group(2).replace("O", "0").replace("Q", "0")
        return t
    return " & ".join(fix(p.strip()) for p in tok.split("&"))


def parse_ew_precinct_text(pages, year, etype, fname, extraction):
    rows = []
    unparsed = []
    suppressed_precincts = set()
    blockcheck = []     # (precinct, contest, parsed_sum, printed_total) mismatches
    for pg in pages:
        lines = [ln.rstrip() for ln in pg.splitlines()]
        lines = [ln for ln in lines if ln.strip()]
        if not lines:
            continue
        # drop report header/footer lines
        body = []
        for ln in lines:
            s = " ".join(ln.split())
            if re.search(r"OFFICIAL RESULTS|Utah County,? Utah\s*$|^Utah County$|^UTAH$|"
                         r"Report generated|Precinct Summary -|^Utah County Municipal|"
                         r"Page \d+ of|Election Summary -|^\d{4} (Municipal|Primary|General)|"
                         r"^General Election$|^November|^June|^September|^Summary Results Report", s):
                continue
            body.append(ln)
        precinct, reg, cast = None, "", ""
        contest, vote_for = None, ""
        in_block = False        # True once inside a contest's candidate table
        block_sum, block_n = 0, 0
        cand_re = OCR_CAND_RE if extraction == "pdf_ocr" else CAND_RE
        i = 0
        while i < len(body):
            s = " ".join(body[i].split())
            up = s.upper()
            if up in ("STATISTICS", "TOTAL", "VOTE %", "TOTAL VOTE %", "TOTAL VOTE%"):
                if contest:
                    in_block = True     # the table header row opens the block
                i += 1
                continue
            m = re.match(r"^Registered Voters(?: - Total)?\s+([\d,]+)$", s)
            if m:
                reg = m.group(1).replace(",", ""); i += 1; continue
            m = re.match(r"^Ballots Cast(?: - Total)?\s+([\d,]+)$", s)
            if m:
                cast = m.group(1).replace(",", ""); i += 1; continue
            m = re.match(r"^(?:Total Votes Cast|Contest Totals)\s+([\d,]+)", s)
            if m or re.match(r"^Total Votes Cast\b|^Contest Totals\b", s):
                # block self-check: printed contest total vs sum of parsed lines
                if m and in_block and block_n:
                    printed = int(m.group(1).replace(",", ""))
                    if printed != block_sum:
                        blockcheck.append((precinct or "?", contest,
                                           block_sum, printed))
                contest, vote_for, in_block = None, "", False
                block_sum, block_n = 0, 0
                i += 1
                continue
            if STAT_RE.match(s) or PCT_ONLY_RE.match(s):
                i += 1
                continue
            m = re.match(r"^Vote For\s*(\d+)$", s)
            if m:
                vote_for = m.group(1)
                in_block = True
                i += 1
                continue
            m = re.match(r"^([0-9A-Z]{2,10})\s+Suppressed$", s, re.I)
            if m:
                # whole-precinct small-count suppression (2026 form): the county
                # prints the precinct code + 'Suppressed' in place of its results.
                suppressed_precincts.add(m.group(1))
                i += 1
                continue
            up_p = up.replace("$", "S") if extraction == "pdf_ocr" else up
            if precinct is None and PRECINCT_RE.match(up_p.replace(" ", "")) and len(s) <= 30 \
               and not re.match(r"^(FOR|AGAINST)$", up_p):
                precinct = _fix_ocr_precinct(up_p) if extraction == "pdf_ocr" else s
                i += 1
                continue
            cm = cand_re.match(body[i])
            if cm and contest and in_block:
                cand = " ".join(cm.group(1).split())
                if extraction == "pdf_ocr":
                    cand = _fix_ocr_candidate(cand)
                if cand.upper() in PSEUDO_CANDIDATES:
                    if cand.upper() in ("OVERVOTES", "UNDERVOTES", "OVER VOTES", "UNDER VOTES"):
                        v = int(cm.group(2).replace(",", ""))
                        emit(rows, year, etype, fname, "", contest, vote_for, precinct or "",
                             cand, v, False, cast, reg,
                             grain="precinct", extraction=extraction)
                        block_sum += v   # 'Contest Totals' includes over/under votes
                        block_n += 1
                    i += 1
                    continue
                if STAT_RE.match(cand):
                    i += 1
                    continue
                v = int(cm.group(2).replace(",", ""))
                emit(rows, year, etype, fname, "", contest, vote_for, precinct or "",
                     cand, v, False, cast, reg,
                     grain="precinct", extraction=extraction)
                block_sum += v
                block_n += 1
                i += 1
                continue
            # otherwise: a contest-name line (may wrap; join while next line is
            # neither Vote For / candidate / stat)
            if re.match(r"^[A-Za-z]", s) and not (in_block and cand_re.match(body[i])):
                contest, vote_for, in_block = s, "", False
                block_sum, block_n = 0, 0
                j = i + 1
                while j < len(body):
                    ns = " ".join(body[j].split())
                    if re.match(r"^Vote For\s*(\d+)$", ns) or STAT_RE.match(ns) \
                       or PCT_ONLY_RE.match(ns) or ns.upper() in ("TOTAL", "TOTAL VOTE %", "TOTAL VOTE%") \
                       or cand_re.match(body[j]):
                        break
                    contest += " " + ns
                    j += 1
                i = j
                continue
            unparsed.append(s)
            i += 1
    return rows, unparsed, suppressed_precincts, blockcheck


# ---------------------------------------------------------------------------
# Parser 4: Electionware countywide summary text (2022 primary, 2023 general;
# reconciliation for most years). Same grammar, no precinct.
# RCV final-round blocks ("One Seat"/"Two Seats" instead of "Vote For N") are
# captured separately — they are FINAL-ROUND tabulations, not first choices.
# ---------------------------------------------------------------------------
def parse_ew_summary_text(txt, year, etype, fname):
    lines = [ln for ln in txt.splitlines()]
    body = []
    for ln in lines:
        s = " ".join(ln.split())
        if not s:
            continue
        if re.search(r"OFFICIAL RESULTS|Utah County,? Utah\s*$|^Utah County$|Report generated|Page \d+ of|"
                     r"Election Summary|Results Summary -|Countywide Summary|^\d{4} .*(Municipal|General|Primary)|"
                     r"^November|^June|^September|^STATISTICS$|^TOTAL$|^Registered Voters|^Ballots Cast|^Voter Turnout", s):
            continue
        body.append(ln)
    results = []          # (contest, vote_for, candidate, votes)
    rcv_finals = []       # (contest, seats_label, candidate, votes, pct)
    contest, vote_for, is_rcv_final, seats_label = None, "", False, ""
    for ln in body:
        s = " ".join(ln.split())
        m = re.match(r"^Vote For (\d+)$", s)
        if m:
            vote_for, is_rcv_final = m.group(1), False
            continue
        m = re.match(r"^(One|Two|Three|Four) Seats?$", s)
        if m:
            is_rcv_final, seats_label = True, s
            vote_for = ""
            continue
        if s in ("TOTAL VOTE %", "VOTE %", "TOTAL"):
            continue
        cm = CAND_RE.match(ln)
        if cm and contest:
            cand = " ".join(cm.group(1).split())
            if cand.startswith("Total Votes") or cand.upper() in PSEUDO_CANDIDATES:
                continue
            v = int(cm.group(2).replace(",", ""))
            if is_rcv_final:
                rcv_finals.append((contest, seats_label, cand, v, cm.group(3) or ""))
            else:
                results.append((contest, vote_for, cand, v))
            continue
        if re.match(r"^[A-Za-z]", s) and not CAND_RE.match(ln):
            # contest names can wrap onto a following short line; EW summaries in
            # this archive keep them on one line — treat each such line as a new
            # contest header.
            contest, vote_for, is_rcv_final = s, "", False
    return results, rcv_finals


# ---------------------------------------------------------------------------
# Parser 5: EVS 2017-style summary ("Contest (Vote for N)" / Times Cast /
# Candidate Party Total).
# ---------------------------------------------------------------------------
def parse_evs_summary_text(txt, year, etype, fname):
    results = []
    contest, vote_for = None, ""
    frag = ""          # candidate-name fragment from a wrapped line
    lines = txt.splitlines()
    for li, ln in enumerate(lines):
        s = " ".join(ln.split())
        if not s:
            continue
        # wrapped candidate name: "<NAME-FRAG>" / "   PTY  1,234" / "<NAME-FRAG-2>"
        m = re.match(r"^\s{12,}([A-Z]{2,4}|WRITE-IN)\s\s+([\d,]+)\s*$", ln)
        if m and contest and frag:
            cand = frag
            nxt = " ".join(lines[li + 1].split()) if li + 1 < len(lines) else ""
            if nxt and re.match(r"^[A-Z][A-Z .'-]+$", nxt) and len(nxt) < 30 \
               and not re.search(r"\(Vote for", nxt):
                cand = cand + " " + nxt
            results.append((contest, vote_for, cand,
                            int(m.group(2).replace(",", "")), m.group(1)))
            frag = ""
            continue
        m = re.match(r"^(.*?)\s*\(Vote for (\d+)\)$", s)
        if m:
            contest, vote_for = m.group(1), m.group(2)
            continue
        # 2018-primary form: "US SENATE (REP)" — no Vote-for clause
        if re.match(r"^[A-Z][A-Z0-9 ./#&'-]*\([A-Z]{2,4}\)$", s):
            contest, vote_for = s, ""
            continue
        if re.match(r"^[A-Z]{2,4}$", s):    # standalone party sub-header line
            continue
        if re.search(r"OFFICIAL RESULTS|Election Summary Report|Utah County|Summary for:|"
                     r"^Registered Voters|^Ballots Cast|^Times Cast|^Candidate|^Total$|^Page: \d+|^Total Votes",
                     s):
            continue
        m = re.match(r"^(.+?)\s\s+([A-Z]{2,10}|WRITE-IN)\s\s+([\d,]+)$", ln.strip())
        if m and contest:
            results.append((contest, vote_for, " ".join(m.group(1).split()),
                            int(m.group(3).replace(",", "")), m.group(2)))
            continue
        m = re.match(r"^(.+?)\s\s+([\d,]+)$", ln.strip())
        if m and contest:
            cand = " ".join(m.group(1).split())
            if cand.startswith("Total Votes") or re.match(r"^[\d,/.%\s]+$", cand):
                continue
            results.append((contest, vote_for, cand, int(m.group(2).replace(",", "")), ""))
            continue
        if contest and re.match(r"^[A-Z][A-Z .'-]+$", s) and len(s) < 40:
            frag = s        # possible first half of a wrapped candidate name
    return results


# ---------------------------------------------------------------------------
# Parser 6: GEMS 2016 summary (two newspaper columns).
# ---------------------------------------------------------------------------
def parse_gems_summary_text(txt, year, etype, fname):
    pages = txt.split("\f")
    results = []
    for pg in pages:
        lines = pg.splitlines()
        if not lines:
            continue
        # find the two column x-offsets from 'Number of Precincts' occurrences
        offs = set()
        for ln in lines:
            for m in re.finditer(r"Number of Precincts", ln):
                offs.add(m.start())
        offs = sorted(offs)
        cut = None
        if len(offs) >= 2:
            cut = offs[-1] - 2
        halves = []
        if cut:
            halves = [[ln[:cut] for ln in lines], [ln[cut:] for ln in lines]]
        else:
            halves = [lines]
        for half in halves:
            contest, vote_for = None, ""
            prev_contest = None
            for ln in half:
                s = " ".join(ln.split())
                if not s:
                    continue
                if re.search(r"Election Summary Report|Utah County|Summary For|Official Canvass|"
                             r"Registered Voters \d|Date:|Time:|Page:|Num\. Report", s):
                    continue
                if re.match(r"^(Number of Precincts|Precincts Reporting|Times Counted|Total Votes)", s):
                    continue
                m = re.match(r"^(?:([A-Z]{2,4})\s\s+)?([\d,]+)\s+[\d.]+%$", ln.strip())
                if m and prev_contest is not None:
                    # wrapped candidate name: the name line was just consumed as a
                    # "contest" header, and this line holds only party+votes
                    # (GEMS wraps long names, e.g. STRAIGHT PARTY's
                    # 'INDEPENDENT AMERICAN PARTY').
                    results.append((prev_contest, "", contest,
                                    int(m.group(2).replace(",", "")), m.group(1) or ""))
                    contest, prev_contest = prev_contest, None
                    continue
                m = re.match(r"^(.+?)\s\s+(?:([A-Z]{2,4})\s+)?([\d,]+)\s+[\d.]+%$", ln.strip())
                if m and contest:
                    results.append((contest, vote_for, " ".join(m.group(1).split()),
                                    int(m.group(3).replace(",", "")), m.group(2) or ""))
                    prev_contest = None
                    continue
                if re.match(r"^[A-Z0-9][A-Z0-9 #/&.'()-]+$", s) and not re.match(r"^[\d,\s%]+$", s):
                    prev_contest = contest if contest else None
                    contest, vote_for = s, ""
    return results


# ---------------------------------------------------------------------------
# by-contest derivation
# ---------------------------------------------------------------------------
HELD = [
    ("lehi",     [r"\bLEHI\b"]),
    ("provo",    [r"\bPROVO\b"]),
    ("orem",     [r"\bOREM\b"]),
    ("vineyard", [r"\bVINEYARD\b"]),
]
DISTRICT_BODY_RE = re.compile(
    r"TRUSTEE|SCHOOL|IMPROVEMENT|SEWER|RECREATION|SERVICE AREA|WATER|IRRIGATION|"
    r"\bFIRE\b|CEMETERY|SPECIAL DISTRICT|METRO", re.I)
RCV_RANK_RE = re.compile(r"\s(\d+)(?:ST|ND|RD|TH) CHOICE$", re.I)
COUNTY_OFFICE_RE = re.compile(
    r"^(?:REP |DEM )?(?:REPUBLICAN FOR |DEMOCRATIC FOR )?(?:UTAH )?COUNTY "
    r"(COMM(?:ISSION(?:ER)?)?\.? ?(?:SEAT )?[A-C]\b|CLERK ?/ ?AUDITOR|CLERK|AUDITOR|"
    r"ATTORNEY|SHERIFF|ASSESSOR|RECORDER|SURVEYOR|TREASURER)", re.I)
MUNI_TOKENS = [
    "ALPINE", "AMERICAN FORK", "CEDAR FORT", "CEDAR HILLS", "EAGLE MOUNTAIN",
    "ELK RIDGE", "FAIRFIELD", "GENOLA", "GOSHEN", "HIGHLAND", "LEHI", "LINDON",
    "MAPLETON", "OREM", "PAYSON", "PLEASANT GROVE", "PROVO", "SALEM", "SANTAQUIN",
    "SARATOGA SPRINGS", "SPANISH FORK", "SPRING LAKE", "SPRINGVILLE", "VINEYARD",
    "WOODLAND HILLS",
]


def parse_contest(contest):
    """Return (kind, jurisdiction_slug, office, district, rcv_rank) —
    kind in {'municipal','county',''}; rcv_rank = int rank for RCV rank-position
    contests (1 = first choice) or 0 for non-RCV."""
    up = " ".join(contest.upper().split())
    rank = 0
    m = RCV_RANK_RE.search(up)
    if m:
        rank = int(m.group(1))
        up = up[:m.start()].strip()
    m = COUNTY_OFFICE_RE.match(up)
    if m:
        off = m.group(1).upper()
        if off.startswith("COMM"):
            return "county", "utah_county", "County Commission", "Seat " + off.rstrip()[-1], rank
        off = re.sub(r"\s*/\s*", "/", off)
        return "county", "utah_county", "County " + off.title(), "", rank
    if DISTRICT_BODY_RE.search(up):
        return "", "", "", "", rank
    if re.search(r"PROPOSITION|PROPOSTION|\bBOND\b|\bTAX\b|AMENDMENT|OPINION QUESTION|"
                 r"RETENTION|JUDICIAL|COURT|STRAIGHT PARTY", up):
        return "", "", "", "", rank
    city = ""
    for tok in MUNI_TOKENS:
        if re.search(r"\b" + tok + r"\b", up):
            city = tok
            break
    if not city:
        return "", "", "", "", rank
    if not re.search(r"COUNCIL|MAYOR", up):
        return "", "", "", "", rank
    juris = ""
    for slug, pats in HELD:
        if any(re.search(p, up) for p in pats):
            juris = slug
            break
    if "MAYOR" in up:
        return "municipal", juris, "Mayor", "", rank
    # council district / seat / term qualifier
    district = ""
    m = re.search(r"DISTRICT\s*#?\s*(\d+)", up)
    if m:
        district = m.group(1)
    else:
        m = re.search(r"CITY ?WIDE\s*(\d+|I{1,3})", up)
        if m:
            n = m.group(1)
            district = "City Wide " + str({"I": 1, "II": 2, "III": 3}.get(n, n))
        else:
            m = re.search(r"SEAT\s*([A-E])\b", up)
            if m:
                district = "Seat " + m.group(1)
            else:
                m = re.search(r"\((TWO|2|FOUR|4)\s*YEAR\)", up)
                if m:
                    district = {"2": "2-Year", "TWO": "2-Year",
                                "4": "4-Year", "FOUR": "4-Year"}[m.group(1)]
    return "municipal", juris, "Council", district, rank


def contest_party(contest):
    """Party embedded in a partisan-primary contest name ('REP Republican for …',
    'Republican for …', 'DEM US SENATE DEM')."""
    up = " ".join(contest.upper().split())
    m = re.match(r"^(REP|DEM)\b", up)
    if m:
        return m.group(1)
    if up.startswith("REPUBLICAN FOR"):
        return "REP"
    if up.startswith("DEMOCRATIC FOR"):
        return "DEM"
    return ""


def party_of(candidate, long_party):
    if long_party and long_party in PARTY_TOKENS and long_party != "NON":
        return long_party
    m = re.search(r"\(([A-Z]{2,4})\)\s*$", candidate.strip())
    if m and m.group(1) in PARTY_TOKENS:
        return m.group(1)
    m = re.match(r"^(" + "|".join(sorted(PARTY_TOKENS)) + r")\s+\S", candidate.strip())
    if m:
        return m.group(1)
    return ""


def load_rcv_registry():
    path = os.path.join(RCV_DIR, "rcv_contests.csv")
    reg = []
    if os.path.exists(path):
        with open(path, newline="", encoding="utf-8") as f:
            reg = list(csv.DictReader(f))
    return reg


def flatten_rcv_rounds():
    """rcv/rounds/*.json (rcvis.com embedded blobs, archived) -> rounds CSVs."""
    for jf in sorted(glob.glob(os.path.join(RCV_DIR, "rounds", "*.json"))):
        d = json.load(open(jf))
        out = jf[:-5] + ".csv"
        rows = []
        elected_by_round, elim_by_round = {}, {}
        for r in d.get("results", []):
            for tr in r.get("tallyResults", []):
                if "elected" in tr:
                    elected_by_round.setdefault(r["round"], []).append(tr["elected"])
                if "eliminated" in tr:
                    elim_by_round.setdefault(r["round"], []).append(tr["eliminated"])
        for r in d.get("results", []):
            for cand, v in sorted(r["tally"].items(), key=lambda kv: -float(kv[1])):
                status = ""
                if cand in elected_by_round.get(r["round"], []):
                    status = "elected"
                elif cand in elim_by_round.get(r["round"], []):
                    status = "eliminated"
                rows.append({"round": r["round"], "candidate": cand,
                             "votes": int(float(v)), "status": status})
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["round", "candidate", "votes", "status"])
            w.writeheader()
            w.writerows(rows)


def main():
    long_rows = []
    official_totals = {}     # (year, etype, contest_normalized) -> {candidate: votes}
    seats_by_contest = {}    # (year, etype, normalized contest) -> vote_for
    recon_report = []

    def norm(c):
        return " ".join(c.upper().split())

    for year, etype, fname, parser in PARSED_SOURCES:
        path = os.path.join(RAW, fname)
        if parser == "wide_csv":
            rows, totals = parse_wide_csv(path, year, etype)
            long_rows.extend(rows)
            agg = defaultdict(dict)
            for (contest, cand), v in totals.items():
                agg[norm(contest)][cand] = v
            for c, d in agg.items():
                official_totals[(year, etype, c)] = d
        elif parser == "xlsx_sovc":
            rows, totals = parse_xlsx_sovc(path, year, etype)
            long_rows.extend(rows)
            agg = defaultdict(dict)
            for (contest, cand), v in totals.items():
                agg[norm(contest)][cand] = v
            for c, d in agg.items():
                official_totals[(year, etype, c)] = d
            for r in rows:
                seats_by_contest.setdefault((year, etype, norm(r["contest"])), r["vote_for"])
        elif parser == "ew_precinct":
            txt = open(os.path.join(TEXT, fname[:-4] + ".txt"), encoding="utf-8").read()
            rows, unparsed, supp, blockcheck = parse_ew_precinct_text(txt.split("\f"), year, etype, fname, "pdf_text")
            if blockcheck:
                recon_report.append(f"{year} {etype}: {len(blockcheck)} contest blocks fail the printed-total "
                                    f"self-check: {blockcheck[:5]}")
            long_rows.extend(rows)
            for r in rows:
                if r["vote_for"]:
                    seats_by_contest.setdefault((year, etype, norm(r["contest"])), r["vote_for"])
            if supp:
                recon_report.append(f"{year} {etype}: {len(supp)} whole precincts printed as 'Suppressed' "
                                    f"(small-count privacy): {sorted(supp)}")
            if unparsed:
                recon_report.append(f"{year} {etype}: {len(unparsed)} unparsed precinct-PDF lines (sample: {unparsed[:3]})")
        elif parser == "ocr_precinct":
            pages = []
            for tf in sorted(glob.glob(os.path.join(OCR19, "*.txt"))):
                pages.append(open(tf, encoding="utf-8").read())
            if not pages:
                recon_report.append(f"{year} {etype}: OCR pages missing — precinct grain skipped")
                continue
            rows, unparsed, _, blockcheck = parse_ew_precinct_text(pages, year, etype, fname, "pdf_ocr")
            # documented, visually-verified OCR corrections (see OCR_PATCHES_2019G)
            for (prec, contest, cand), v in OCR_PATCHES_2019G.items():
                if any(r["precinct"] == prec and r["contest"] == contest
                       and r["candidate"] == cand for r in rows):
                    continue        # already recovered by a better OCR pass
                ctx = next((r for r in rows if r["precinct"] == prec), {})
                emit(rows, year, etype, fname, "", contest, "1", prec, cand, v, False,
                     ctx.get("times_cast", ""), ctx.get("registered_voters", ""),
                     grain="precinct", extraction="pdf_ocr+visual")
            long_rows.extend(rows)
            if blockcheck:
                recon_report.append(f"{year} {etype} (OCR): {len(blockcheck)} contest blocks fail the "
                                    f"printed-total self-check: {blockcheck[:8]}")
            for r in rows:
                if r["vote_for"]:
                    seats_by_contest.setdefault((year, etype, norm(r["contest"])), r["vote_for"])
            if unparsed:
                recon_report.append(f"{year} {etype} (OCR): {len(unparsed)} unparsed lines (sample: {unparsed[:3]})")
        elif parser == "ew_summary":
            txt = open(os.path.join(TEXT, fname[:-4] + ".txt"), encoding="utf-8").read()
            results, rcv_finals = parse_ew_summary_text(txt, year, etype, fname)
            for contest, vote_for, cand, v in results:
                emit(long_rows, year, etype, fname, "", contest, vote_for, "", cand, v,
                     False, grain="countywide", extraction="pdf_text")
                seats_by_contest.setdefault((year, etype, norm(contest)), vote_for)
        elif parser == "evs_summary":
            txt = open(os.path.join(TEXT, fname[:-4] + ".txt"), encoding="utf-8").read()
            results = parse_evs_summary_text(txt, year, etype, fname)
            for contest, vote_for, cand, v, pty in results:
                emit(long_rows, year, etype, fname, "", contest, vote_for, "", cand, v,
                     False, party=pty, grain="countywide", extraction="pdf_text")
                seats_by_contest.setdefault((year, etype, norm(contest)), vote_for)
        elif parser == "gems_summary":
            txt = open(os.path.join(TEXT, fname[:-4] + ".txt"), encoding="utf-8").read()
            results = parse_gems_summary_text(txt, year, etype, fname)
            for contest, vote_for, cand, v, pty in results:
                emit(long_rows, year, etype, fname, "", contest, vote_for, "", cand, v,
                     False, party=pty, grain="countywide", extraction="pdf_text")

    # drop countywide rows for (year, etype, contest) already covered at precinct grain
    prec_keys = {(r["year"], r["election_type"], norm(r["contest"]))
                 for r in long_rows if r["grain"] == "precinct"}
    long_rows = [r for r in long_rows
                 if r["grain"] == "precinct"
                 or (r["year"], r["election_type"], norm(r["contest"])) not in prec_keys]

    long_rows.sort(key=lambda r: (r["year"], r["election_type"], r["contest"],
                                  r["precinct"], r["candidate"]))
    cols = ["year", "election_type", "source_file", "sheet", "contest", "vote_for",
            "precinct", "candidate", "votes", "suppressed", "vote_method",
            "times_cast", "registered_voters", "party", "grain", "extraction"]
    with open(LONG_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in long_rows:
            r2 = dict(r)
            r2["suppressed"] = "True" if r["suppressed"] else "False"
            w.writerow(r2)
    print(f"Wrote {LONG_OUT}: {len(long_rows)} rows")

    # ------------------------------------------------------------------
    # reconciliation: precinct sums vs COUNTY TOTALS (matrix rollup)
    # ------------------------------------------------------------------
    sums = defaultdict(float)
    for r in long_rows:
        if r["grain"] == "precinct" and r["votes"] != "":
            sums[(r["year"], r["election_type"], norm(r["contest"]), r["candidate"])] += r["votes"]
    n_exact = n_under = n_over = 0
    for (year, etype, c), cands in official_totals.items():
        for cand, tot in cands.items():
            s = sums.get((year, etype, c, cand))
            if s is None:
                continue
            if int(s) == tot:
                n_exact += 1
            elif int(s) < tot:
                n_under += 1
            else:
                n_over += 1
                recon_report.append(f"OVER: {year} {etype} {c} / {cand}: precinct-sum {int(s)} > official {tot}")
    print(f"COUNTY-TOTALS reconciliation: exact={n_exact} under(=suppressed remainder)={n_under} OVER={n_over}")
    for line in recon_report:
        print("  NOTE:", line)

    # ------------------------------------------------------------------
    # reconciliation vs the county's own countywide summary PDFs
    # (per-candidate: summary total vs official COUNTY TOTALS where present,
    # else vs precinct sum; a precinct-sum shortfall vs the summary is the
    # suppressed remainder unless COUNTY TOTALS covers it)
    # ------------------------------------------------------------------
    def contest_candidates(year, etype):
        d = defaultdict(set)
        for r in long_rows:
            if r["year"] == year and r["election_type"] == etype:
                d[norm(r["contest"])].add(r["candidate"])
        return d

    print("SUMMARY-PDF reconciliation (per candidate):")
    for year, etype, tfn, fmt, official in RECON_SUMMARIES:
        tp = os.path.join(TEXT, tfn)
        if not os.path.exists(tp):
            continue
        txt = open(tp, encoding="utf-8").read()
        if fmt == "evs":
            results = [(c, vf, cand, v) for c, vf, cand, v, _ in
                       parse_evs_summary_text(txt, year, etype, tfn)]
        else:
            results, rcv_finals = parse_ew_summary_text(txt, year, etype, tfn)
        have = contest_candidates(year, etype)
        n_ok = n_diff = n_low = n_missing = 0
        diffs = []
        for contest, vote_for, cand, v in results:
            c = norm(contest)
            if cand.upper() in PSEUDO_CANDIDATES:
                continue
            if c not in have:
                # RCV rank-form contests appear in the SOVC as '<contest> 1st Choice'
                if (c + " 1ST CHOICE") in have:
                    c = c + " 1ST CHOICE"
                else:
                    n_missing += 1
                    continue
            off = official_totals.get((year, etype, c), {})
            mine = off.get(cand)
            if mine is None:
                s = sums.get((year, etype, c, cand))
                mine = int(s) if s is not None else None
            if mine is None:
                # summary write-in itemizations / wrapped-name variants: try
                # prefix containment before declaring missing
                pref = [k for k in official_totals.get((year, etype, c), {})
                        if k.upper().startswith(cand.upper())]
                if len(pref) == 1:
                    mine = official_totals[(year, etype, c)][pref[0]]
            if mine is None:
                n_missing += 1
                diffs.append(f"MISSING {c} / {cand} (summary {v})")
            elif mine == v:
                n_ok += 1
            elif not official and mine >= v:
                n_low += 1     # interim/unofficial report running low — expected
            else:
                n_diff += 1
                diffs.append(f"DIFF {c} / {cand}: mine {mine} vs summary {v}")
        tag = "" if official else " (UNOFFICIAL interim report — sovc>report expected)"
        print(f"  {year} {etype} [{tfn[:40]}]: ok={n_ok} diff={n_diff} "
              f"interim-low={n_low} missing={n_missing}{tag}")
        # The 2019 general precinct grain is OCR from a scan MISSING 5 of 261
        # report pages — the certified countywide summary is the authoritative
        # total for that election, so its per-candidate numbers become the
        # official totals the by-contest layer prefers over precinct sums.
        if official and (year, etype) == (2019, "municipal general"):
            for contest, vote_for, cand, v in results:
                c = norm(contest)
                official_totals.setdefault((year, etype, c), {}).setdefault(cand, v)
                if vote_for:
                    seats_by_contest.setdefault((year, etype, c), vote_for)
        for d in diffs[:12]:
            print("    ", d)
        if len(diffs) > 12:
            print(f"     ... {len(diffs)-12} more")

    # ------------------------------------------------------------------
    # by-contest
    # ------------------------------------------------------------------
    flatten_rcv_rounds()
    rcv_reg = load_rcv_registry()
    rcv_final = {}
    for row in rcv_reg:
        key = (int(row["year"]), row["election_type"], norm(row["sovc_first_choice_contest"] or row["contest"]))
        rcv_final[key] = row["final_winners"]

    agg = defaultdict(lambda: {"votes": 0.0, "precincts": set(), "source": "",
                               "suppressed": False, "party": ""})
    contest_meta = {}
    for r in long_rows:
        kind, juris, office, district, rank = parse_contest(r["contest"])
        if not kind or rank > 1:
            continue
        if r["candidate"].upper() in PSEUDO_CANDIDATES:
            continue
        key = (r["year"], r["election_type"], r["contest"], r["candidate"])
        a = agg[key]
        if r["votes"] != "":
            a["votes"] += r["votes"]
        if r["suppressed"]:
            a["suppressed"] = True
        if r["grain"] == "precinct" and r["precinct"]:
            a["precincts"].add(r["precinct"])
        a["source"] = r["source_file"]
        if not a["party"]:
            a["party"] = party_of(r["candidate"], r["party"]) or contest_party(r["contest"])
        contest_meta[(r["year"], r["election_type"], r["contest"])] = (kind, juris, office, district, rank)

    # RCV-only contests (absent from the county SOVC entirely — 2023 general +
    # 2023 Lehi primary): first-round tallies from the county-linked rcvis
    # tabulations, marked rcv, n_precincts=0.
    for row in rcv_reg:
        if row.get("in_sovc") != "absent":
            continue
        jf = os.path.join(RCV_DIR, "rounds", row["primary_slug"] + ".json")
        if not os.path.exists(jf):
            continue
        d = json.load(open(jf))
        r1 = d["results"][0]["tally"]
        year, etype, contest = int(row["year"]), row["election_type"], row["contest"]
        for cand, v in r1.items():
            if cand == "Inactive Ballots":
                continue
            key = (year, etype, contest, cand)
            agg[key]["votes"] = float(v)
            agg[key]["source"] = "rcvis.com/v/" + row["primary_slug"]
        kind, juris, office, district, _ = parse_contest(contest)
        contest_meta[(year, etype, contest)] = (kind, juris, office, district, 1)

    by_contest = defaultdict(list)
    for (year, etype, contest, cand), a in agg.items():
        by_contest[(year, etype, contest)].append((cand, a))
    rows = []
    for (year, etype, contest), cands in sorted(by_contest.items()):
        kind, juris, office, district, rank = contest_meta[(year, etype, contest)]
        is_rcv = rank == 1 and (RCV_RANK_RE.search(contest.upper()) is not None
                                or (year, etype, norm(contest)) in rcv_final)
        cands.sort(key=lambda ca: ca[1]["votes"], reverse=True)
        # official totals (county rollup) where present
        off = official_totals.get((year, etype, norm(contest)), {})
        seats = seats_by_contest.get((year, etype, norm(contest)), "")
        for rk, (cand, a) in enumerate(cands, start=1):
            v_official = off.get(cand)
            rows.append({
                "year": year, "election_type": etype, "contest": contest,
                "jurisdiction_slug": juris, "office": office, "district": district,
                "seats": seats, "candidate": cand, "party": a["party"],
                "votes": v_official if v_official is not None else int(round(a["votes"])),
                "rank_in_contest": rk,
                "n_precincts": len(a["precincts"]),
                "suppressed": "true" if (a["suppressed"] and v_official is None) else "false",
                "source_file": a["source"],
                "rcv": "true" if is_rcv else "false",
                "rcv_final_winner": rcv_final.get((year, etype, norm(contest)), "") if is_rcv else "",
                "official_total": "true" if v_official is not None else "false",
            })
    cols = ["year", "election_type", "contest", "jurisdiction_slug", "office",
            "district", "seats", "candidate", "party", "votes", "rank_in_contest",
            "n_precincts", "suppressed", "source_file", "rcv", "rcv_final_winner",
            "official_total"]
    rows.sort(key=lambda x: (x["year"], x["election_type"], x["jurisdiction_slug"],
                             x["office"], x["district"], x["rank_in_contest"]))
    with open(BYC_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)
    juris_n = defaultdict(int)
    for r in rows:
        juris_n[r["jurisdiction_slug"] or "(other-muni)"] += 1
    print(f"Wrote {BYC_OUT}: {len(rows)} contest×candidate rows, {len(by_contest)} contests")
    print("  rows by jurisdiction:", dict(sorted(juris_n.items())))


if __name__ == "__main__":
    main()
